"""
Transaction Admin — FastAPI backend v2
PostgreSQL + Summary + Reconciliation (Google Sheets)
"""

import json
import math
import os
import re
import time
from io import BytesIO

import numpy as np
import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse

# ── optional Google Sheets ────────────────────────────────────────────────────
try:
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build
    HAS_GOOGLE = True
except ImportError:
    HAS_GOOGLE = False

# =============================================================================
# CONFIG
# =============================================================================

SPREADSHEET_ID     = os.getenv("SPREADSHEET_ID", "")
RECON_SHEET_NAME   = os.getenv("RECON_SHEET_NAME", "1. Reconciliation")
SCOPES             = ["https://www.googleapis.com/auth/spreadsheets"]
SHEETS_WRITE_CHUNK = 2000
HEADER_ROWS        = 2
DATA_START_ROW     = HEADER_ROWS + 1

EXCLUDED_MERCHANTS = {
    "sognolab", "cratecracker", "fiftytemps", "asquad",
    "profit bridge ltd", "ideayard", "skinsbo"
}
CARD_PAY_TYPES = {"visa", "mastercard", "maestro"}
OB_PAY_TYPES   = {"open-banking", "banks/germany"}
SHOW_STATUSES  = {"success", "decline", "processing"}

# код метода Apple Pay в новой колонке "Payment method"
APPLE_PAY_METHOD_CODE = "69"

COLUMN_MAP = {
    "provider payment id": ["acquirer_id / provider_payment_id", "provider_payment_id / acquirer_id", "provider payment id", "provider_payment_id"],
    "currency": ["currency", "currency / currency"],
    "transaction status": ["transaction status", "operation_status", "status"],
    "external id payment id": ["external_id / payment_id", "external_id", "payment_id"],
    "operation type": ["operation type", "operation_type", "type"],
    "amount": ["amount", "amount / amount"],
    "мой выгрузка": ["created_at / operation_created_at", "created_at", "operation_created_at"],
    "merchant name": ["merchant name", "merchant_name"],
    "payment_type_id": ["payment_type_id", "payment_type_id / payment_method_type", "payment_method_type"],
    "id / operation_id": ["id / operation_id", "operation_id", "id"],
    "customer email": ["customer email", "customer_email", "email"],
    "customer name": ["customer name"],
    "match": ["match", "Match"],
    "correct reconcile ?": ["correct reconcile ?", "correct_reconcile"],
    "fraud": ["fraud", "Fraud"],
    "fraud action ?": ["fraud action ?", "fraud_action"],
    "payment method": ["payment method", "payment_method"],
}

# =============================================================================
# APP
# =============================================================================

app = FastAPI(title="Transaction Admin API v2")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def root():
    return {"status": "ok", "version": "2.0"}

@app.get("/app", response_class=HTMLResponse)
def serve_app():
    if os.path.exists("index.html"):
        with open("index.html", "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse(content="<h1>index.html not found</h1>", status_code=404)

# =============================================================================
# HELPERS
# =============================================================================

def norm_key(value):
    return re.sub(r"\s+", " ", (value or "").strip().lower())

def strip_apostrophe(value):
    s = str(value or "").strip()
    return s[1:].lstrip() if s.startswith("'") else s

def read_csv(data: bytes) -> pd.DataFrame:
    df = pd.read_csv(BytesIO(data), sep=";", dtype=str, keep_default_na=False)
    if df.shape[1] <= 1:
        df = pd.read_csv(BytesIO(data), sep=",", dtype=str, keep_default_na=False)
    return df

def to_float(x):
    s = str(x or "").strip().replace(" ", "").replace(",", ".")
    try: return float(s)
    except: return 0.0

def to_amount(x):
    s = strip_apostrophe(x).replace(" ", "")
    if not s: return 0.0
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try: return float(s)
    except: return 0.0

def to_safe_int(x, max_len=15):
    s = strip_apostrophe(x)
    if not s: return ""
    if s.isdigit() and len(s) <= max_len:
        try: return int(s)
        except: pass
    return s

def parse_date_ddmmyyyy(x):
    s = strip_apostrophe(x)
    if not s: return ""
    dt = pd.to_datetime(s, utc=True, errors="coerce")
    return "" if pd.isna(dt) else dt.strftime("%d.%m.%Y")

def fmt(x): return f"{x:,.2f}"
def trx_word(n): return "transaction" if n == 1 else "transactions"

# ИСПРАВЛЕНО: убрана замена названий мерчантов (temgo -> afribaba удалено)
def normalize_merchant(name):
    return (name or "").strip() or "unknown"

def find_col(df, variants):
    index = {norm_key(col): col for col in df.columns}
    for v in variants:
        col = index.get(norm_key(v))
        if col: return col
    return None

# =============================================================================
# SUMMARY
# =============================================================================

def _status_lines(grp, status):
    x = grp[grp["status"] == status]
    if x.empty: return 0, []
    cnt = int(x["cnt"].sum())
    if x["cur"].nunique() == 1:
        return cnt, [f"• {fmt(float(x['amt'].sum()))} {x['cur'].iloc[0].upper()}"]
    return cnt, [f"• {fmt(float(r['amt']))} {r['cur'].upper()}" for _, r in x.iterrows()]

def _merchant_block(grp, op_type, title):
    df = grp[grp["op_type"] == op_type].copy()
    succ_cnt, succ_lines = _status_lines(df, "success")
    decl_cnt, decl_lines = _status_lines(df, "decline")
    proc_cnt, proc_lines = _status_lines(df, "processing")
    total = succ_cnt + decl_cnt + proc_cnt
    if total == 0: return "", succ_cnt, decl_cnt, proc_cnt, 0.0
    conv  = succ_cnt / total * 100
    badge = "✅" if conv >= 50 else "⚠️"
    lines = [title]
    if succ_cnt: lines += [f"successful: {succ_cnt} {trx_word(succ_cnt)}"] + succ_lines
    if decl_cnt: lines += [f"declined: {decl_cnt} {trx_word(decl_cnt)}"] + decl_lines
    if proc_cnt: lines += [f"processing: {proc_cnt} {trx_word(proc_cnt)}"] + proc_lines
    lines.append(f"conversion: {conv:.2f}% successful {badge}")
    return "\n".join(lines), succ_cnt, decl_cnt, proc_cnt, conv

def _total_line(tot, status):
    x = tot[tot["status"] == status]
    if x.empty: return "0"
    if x["cur"].nunique() == 1:
        return f"{fmt(float(x['amt'].sum()))} {x['cur'].iloc[0].lower()}"
    return " / ".join(f"{fmt(float(r['amt']))} {r['cur'].lower()}" for _, r in x.iterrows())

def _build_block_section(data, op_type, title_word):
    """Собирает total-строку + блоки по мерчантам для произвольного датафрейма."""
    if data.empty:
        return None
    tot = data.groupby(["status", "cur"], dropna=False).agg(cnt=("amt", "size"), amt=("amt", "sum")).reset_index()
    total_cnt = int(len(data))
    succ_cnt  = int(tot[tot["status"] == "success"]["cnt"].sum())
    decl_cnt  = int(tot[tot["status"] == "decline"]["cnt"].sum())
    proc_cnt  = int(tot[tot["status"] == "processing"]["cnt"].sum())

    msg = f"total: {total_cnt} {trx_word(total_cnt)}\n"
    msg += f"{succ_cnt} success for {_total_line(tot,'success')}\n"
    msg += f"{decl_cnt} decline for {_total_line(tot,'decline')}\n"
    msg += f"{proc_cnt} processing for {_total_line(tot,'processing')}\n\n"

    grp = data.groupby(["merchant", "op_type", "status", "cur"], dropna=False).agg(cnt=("amt", "size"), amt=("amt", "sum")).reset_index()
    order = grp.groupby("merchant")["cnt"].sum().sort_values(ascending=False).index.tolist()
    blocks = []
    for i, merchant in enumerate(order, start=1):
        block, *_ = _merchant_block(grp[grp["merchant"] == merchant], op_type, title_word)
        if block:
            blocks.append(f"{i} - {merchant}\n\n{block}")
    msg += "\n\n".join(blocks) if blocks else "No blocks found."
    return msg


def build_report(df, op_filter="deposit"):
    required = ["merchant_name", "operation_status", "operation_type", "amount / amount", "currency / currency"]
    payment_col = "payment_type_id / payment_method_type"
    has_pay_col = payment_col in df.columns
    if has_pay_col: required.append(payment_col)
    missing = [c for c in required if c not in df.columns]
    if missing:
        return "❌ Missing columns: " + ", ".join(missing)

    # новая колонка "Payment method" (последняя в файле): 3221=OB, 1=карты, 69=Apple Pay
    method_col = find_col(df, COLUMN_MAP["payment method"])
    has_method_col = method_col is not None

    work = df.copy()
    # ИСПРАВЛЕНО: оставляем оригинальное название мерчанта, только приводим к lower для группировки
    work["merchant"] = work["merchant_name"].astype(str).str.strip().str.lower()
    work["_low"]     = work["merchant_name"].astype(str).str.strip().str.lower()
    work             = work[~work["_low"].isin(EXCLUDED_MERCHANTS)].drop(columns=["_low"])
    work["status"]   = work["operation_status"].astype(str).str.strip().str.lower()
    work["op_type"]  = work["operation_type"].astype(str).str.strip().str.lower()
    work["cur"]      = work["currency / currency"].astype(str).str.strip().str.upper()
    work["amt"]      = work["amount / amount"].apply(to_float)
    work["pay_type"] = work[payment_col].astype(str).str.strip().str.lower() if has_pay_col else ""
    work["pay_method"] = work[method_col].astype(str).map(strip_apostrophe).str.strip() if has_method_col else ""

    label   = "📥 Deposit" if op_filter == "deposit" else "📤 Payout"
    card_op = "sale" if op_filter == "deposit" else "payout"

    # маска Apple Pay по новой колонке "Payment method" == 69
    is_apple = work["pay_method"] == APPLE_PAY_METHOD_CODE

    # ── CARD (исключая Apple Pay) ──
    card = work[
        work["pay_type"].isin(CARD_PAY_TYPES)
        & (work["op_type"] == card_op)
        & work["status"].isin(SHOW_STATUSES)
        & ~is_apple
    ]
    card_msg = f"💳 card: {label} (file)\n\n"
    section = _build_block_section(card, card_op, "deposits:" if op_filter == "deposit" else "payouts:")
    card_msg += section if section is not None else "No card transactions found."

    # ── APPLE PAY (Payment method == 69) ──
    apple = work[
        work["pay_type"].isin(CARD_PAY_TYPES)
        & (work["op_type"] == card_op)
        & work["status"].isin(SHOW_STATUSES)
        & is_apple
    ]
    apple_msg = f"🍏 apple pay: {label} (file)\n\n"
    section = _build_block_section(apple, card_op, "deposits:" if op_filter == "deposit" else "payouts:")
    apple_msg += section if section is not None else "No apple pay transactions found."

    # ── OB ──
    ob_base = work[work["pay_type"].isin(OB_PAY_TYPES)] if has_pay_col and work["pay_type"].isin(OB_PAY_TYPES).any() else work

    if op_filter == "deposit":
        ob_op = "payment confirmation"
        ob_success_proc  = ob_base[(ob_base["op_type"] == "payment confirmation") & ob_base["status"].isin({"success","processing"})]
        ob_decline_pc    = ob_base[(ob_base["op_type"] == "payment confirmation") & (ob_base["status"] == "decline")]
        ob_decline_sale  = ob_base[(ob_base["op_type"] == "sale") & (ob_base["status"] == "decline")]
        ob_decline       = ob_decline_pc if not ob_decline_pc.empty else ob_decline_sale
        ob = pd.concat([ob_success_proc, ob_decline], ignore_index=True)
        ob["op_type"] = ob_op
    else:
        ob_op = "payout"
        ob = ob_base[(ob_base["op_type"] == ob_op) & ob_base["status"].isin(SHOW_STATUSES)].copy()

    ob_msg = f"💰 OB: {label} (file)\n\n"
    if ob.empty:
        ob_msg += "No OB transactions found."
        return (card_msg + "\n\n" + apple_msg + "\n\n" + ob_msg).strip()

    ob_tot    = ob.groupby(["status","cur"], dropna=False).agg(cnt=("amt","size"), amt=("amt","sum")).reset_index()
    total_cnt = int(len(ob))
    succ_cnt  = int(ob_tot[ob_tot["status"]=="success"]["cnt"].sum())
    decl_cnt  = int(ob_tot[ob_tot["status"]=="decline"]["cnt"].sum())
    proc_cnt  = int(ob_tot[ob_tot["status"]=="processing"]["cnt"].sum())
    conv      = succ_cnt / total_cnt * 100 if total_cnt > 0 else 0
    badge     = "✅" if conv >= 50 else "⚠️"
    ob_msg   += f"total: {total_cnt} {trx_word(total_cnt)}\n{succ_cnt} success for {_total_line(ob_tot,'success')}\n{decl_cnt} decline for {_total_line(ob_tot,'decline')}\n{proc_cnt} processing for {_total_line(ob_tot,'processing')}\nconversion: {conv:.2f}% successful {badge}\n\n"

    grp2  = ob.groupby(["merchant","op_type","status","cur"], dropna=False).agg(cnt=("amt","size"), amt=("amt","sum")).reset_index()
    order2 = grp2.groupby("merchant")["cnt"].sum().sort_values(ascending=False).index.tolist()
    blocks2 = []
    for i, merchant in enumerate(order2, start=1):
        block, *_ = _merchant_block(grp2[grp2["merchant"]==merchant], ob_op, "deposits:" if op_filter=="deposit" else "payouts:")
        if block: blocks2.append(f"{i} - {merchant}\n\n{block}")
    ob_msg += "\n\n".join(blocks2) if blocks2 else "No blocks found."

    return (card_msg + "\n\n" + apple_msg + "\n\n" + ob_msg).strip()

# =============================================================================
# GOOGLE SHEETS (RECONCILIATION)
# =============================================================================

def _sheets_service():
    if not HAS_GOOGLE:
        raise RuntimeError("google-api-python-client не установлен")
    token_raw = os.getenv("GOOGLE_TOKEN_JSON", "")
    if not token_raw:
        raise RuntimeError("GOOGLE_TOKEN_JSON не задан")
    creds = Credentials.from_authorized_user_info(json.loads(token_raw), SCOPES)
    return build("sheets", "v4", credentials=creds)

def _get_recon_headers(service):
    result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{RECON_SHEET_NAME}'!A2:ZZ2",
    ).execute()
    rows = result.get("values", [])
    return rows[0] if rows else []

def _clear_data(service):
    service.spreadsheets().values().clear(
        spreadsheetId=SPREADSHEET_ID,
        range=f"'{RECON_SHEET_NAME}'!A{DATA_START_ROW}:ZZ",
    ).execute()

def _col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s or "A"

def _write_chunk(service, rng, chunk, retries=6):
    delay = 10
    for attempt in range(retries):
        try:
            service.spreadsheets().values().update(
                spreadsheetId=SPREADSHEET_ID, range=rng,
                valueInputOption="USER_ENTERED", body={"values": chunk},
            ).execute()
            return
        except Exception as e:
            if any(x in str(e) for x in ("429", "RATE_LIMIT", "Quota")) and attempt < retries - 1:
                time.sleep(delay); delay = min(delay * 2, 120)
            else:
                raise

def _write_all(service, values, start_row, col_count):
    if not values: return
    end_col = _col_letter(col_count)
    idx = 0
    while idx < len(values):
        chunk = values[idx: idx + SHEETS_WRITE_CHUNK]
        r1 = start_row + idx; r2 = r1 + len(chunk) - 1
        rng = f"'{RECON_SHEET_NAME}'!A{r1}:{end_col}{r2}"
        _write_chunk(service, rng, chunk)
        idx += len(chunk)
        if idx < len(values): time.sleep(1.5)

def _prepare_recon(df, headers):
    out = pd.DataFrame(index=df.index)
    first_col   = find_col(df, ["first_name","customer_first_name","firstname","first name"])
    last_col    = find_col(df, ["last_name","customer_last_name","lastname","last name"])
    created_col = find_col(df, ["created_at / operation_created_at","created_at","operation_created_at"])
    amount_col  = find_col(df, ["amount","amount / amount"])
    extid_col   = find_col(df, ["external_id / payment_id","external_id","payment_id"])
    opid_col    = find_col(df, ["id / operation_id","operation_id","id"])

    for h in headers:
        hk = norm_key(h)
        if hk == "merchant name":
            # ИСПРАВЛЕНО: убран вызов normalize_merchant, оставляем оригинальное название
            col = find_col(df, ["merchant_name", "merchant name", "merchant"])
            out[h] = df[col].astype(str).map(strip_apostrophe).str.strip() if col else ""
        elif hk == "customer name":
            first = df[first_col].astype(str).map(strip_apostrophe).str.strip() if first_col else pd.Series("", index=df.index)
            last  = df[last_col].astype(str).map(strip_apostrophe).str.strip()  if last_col  else pd.Series("", index=df.index)
            out[h] = (first + " " + last).str.strip()
        elif hk == norm_key("мой выгрузка"):
            out[h] = df[created_col].apply(parse_date_ddmmyyyy) if created_col else ""
        elif hk == "amount":
            out[h] = df[amount_col].apply(to_amount) if amount_col else 0.0
        elif hk == norm_key("external id payment id"):
            out[h] = df[extid_col].apply(lambda x: to_safe_int(x, 15)) if extid_col else ""
        elif hk == norm_key("id / operation_id"):
            out[h] = df[opid_col].apply(lambda x: to_safe_int(x, 15)) if opid_col else ""
        else:
            variants = COLUMN_MAP.get(hk, [h])
            col = find_col(df, variants)
            out[h] = df[col].astype(str).map(strip_apostrophe) if col else ""
    return out

def _df_to_values(df):
    df = df.replace({np.nan: "", np.inf: "", -np.inf: ""})
    result = []
    for row in df.values.tolist():
        out = []
        for v in row:
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                out.append("")
            else:
                out.append("" if v is None else v)
        result.append(out)
    return result

# =============================================================================
# API ENDPOINTS
# =============================================================================

@app.post("/summary")
async def api_summary(file: UploadFile = File(...), op_type: str = Form("deposit")):
    data = await file.read()
    try:
        df = read_csv(data)
    except Exception as e:
        raise HTTPException(400, f"Не смог прочитать CSV: {e}")
    report = build_report(df, op_filter=op_type)
    return {"report": report, "rows": len(df)}

@app.post("/recon")
async def api_recon(file: UploadFile = File(...)):
    if not SPREADSHEET_ID:
        raise HTTPException(400, "SPREADSHEET_ID не задан")
    data = await file.read()
    try:
        df = read_csv(data)
    except Exception as e:
        raise HTTPException(400, f"Не смог прочитать CSV: {e}")
    try:
        service = _sheets_service()
        headers = _get_recon_headers(service)
        prepared = _prepare_recon(df, headers)
        values = _df_to_values(prepared)
        _clear_data(service)
        _write_all(service, values, DATA_START_ROW, len(headers))
    except Exception as e:
        raise HTTPException(500, str(e))
    return {"ok": True, "rows": len(df)}


# =============================================================================
# NOBIMATIK — XLS bank vs CSV deposit matching
# =============================================================================

def _read_bank(data: bytes) -> pd.DataFrame:
    """Read bank file (XLS or CSV), return only BNK credit rows."""
    try:
        df_raw = pd.read_excel(BytesIO(data), engine="xlrd", header=None, dtype=str)
        header_row = 12
        for i, row in df_raw.iterrows():
            vals = [str(v).strip().lower() for v in row if str(v).strip() not in ("nan","")]
            if any("transaction date" in v or "transaction reg" in v for v in vals):
                header_row = i
                break
        df = pd.read_excel(BytesIO(data), engine="xlrd", header=header_row, dtype=str)
    except Exception:
        for enc in ("utf-8", "windows-1251", "latin-1"):
            try:
                df = pd.read_csv(BytesIO(data), sep=";", dtype=str,
                                 keep_default_na=False, encoding=enc)
                if df.shape[1] <= 1:
                    df = pd.read_csv(BytesIO(data), sep=",", dtype=str,
                                     keep_default_na=False, encoding=enc)
                break
            except Exception:
                continue

    df.columns = [str(c).strip() for c in df.columns]

    def _col(kws):
        for c in df.columns:
            if any(k in c.lower() for k in kws):
                return c
        return None

    credit_col = _col(["credit(c)", "credit"])
    trn_col    = _col(["transaction registration", "registration number"])
    date_col   = _col(["transaction date", "date"])
    remit_col  = _col(["remitter/beneficiary", "remitter", "beneficiary"])

    if not credit_col or not trn_col:
        raise ValueError(f"Не найдены колонки Credit/TRN. Колонки: {list(df.columns)}")

    df["_credit"] = df[credit_col].apply(to_amount)
    df["_bnk"]    = df[trn_col].astype(str).str.strip()
    df["_date"]   = df[date_col].astype(str).str.strip() if date_col else ""
    df["_remit"]  = df[remit_col].astype(str).str.strip() if remit_col else ""

    return df[
        (df["_credit"] > 0.5) &
        (df["_bnk"].str.upper().str.startswith("BNK"))
    ].copy().reset_index(drop=True)


def _read_deposit(data: bytes) -> pd.DataFrame:
    """Read deposit CSV with encoding detection."""
    for enc in ("utf-8", "windows-1251", "latin-1"):
        try:
            df = pd.read_csv(BytesIO(data), sep=";", dtype=str,
                             keep_default_na=False, encoding=enc)
            if df.shape[1] <= 1:
                df = pd.read_csv(BytesIO(data), sep=",", dtype=str,
                                 keep_default_na=False, encoding=enc)
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception:
            continue
    raise ValueError("Не удалось прочитать депозитный CSV")


def build_nobimatik_report(bank_data: bytes, deposit_data: bytes):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    bank_cr = _read_bank(bank_data)

    df_dep = _read_deposit(deposit_data)

    def _col(df, kws):
        for c in df.columns:
            if any(k in c.lower() for k in kws):
                return c
        return None

    acq_col   = _col(df_dep, ["acquirer_id"])
    ext_col   = _col(df_dep, ["external_id / payment_id", "external_id"])
    if not ext_col:
        ext_col = _col(df_dep, ["payment_id"])
    email_col = _col(df_dep, ["customer_email", "email"])
    name_col  = _col(df_dep, ["customer_name", "name"])
    amt_col   = _col(df_dep, ["amount"])
    cur_col   = _col(df_dep, ["currency"])
    dat_col   = _col(df_dep, ["created_at", "operation_created"])
    stat_col  = _col(df_dep, ["operation_status", "status"])
    op_col    = _col(df_dep, ["operation_type", "type"])

    if not acq_col:
        raise ValueError(f"Не найдена колонка acquirer_id. Колонки: {list(df_dep.columns)}")

    mask = pd.Series([True] * len(df_dep), index=df_dep.index)
    if stat_col: mask &= df_dep[stat_col].str.strip().str.lower() == "success"
    if op_col:   mask &= df_dep[op_col].str.strip().str.lower() == "payment confirmation"
    df_dep_f = df_dep[mask].copy()

    bnk_to_dep = {}
    for _, row in df_dep_f.iterrows():
        bnk = str(row[acq_col]).strip()
        if bnk.upper().startswith("BNK"):
            bnk_to_dep[bnk] = row

    bank_cr["_matched"] = bank_cr["_bnk"].isin(bnk_to_dep)
    matched   = bank_cr[bank_cr["_matched"]].reset_index(drop=True)
    not_found = bank_cr[~bank_cr["_matched"]].reset_index(drop=True)

    NAVY = "FF1F3864"; WHITE = "FFFFFFFF"
    GREEN = "FFE2EFDA"; GREEN2 = "FFD0E4C8"
    RED = "FFFCE4D6";   RED2  = "FFFAD7CC"
    FN = "Verdana"

    def hdr(cell, text):
        cell.value = text
        cell.font      = Font(name=FN, bold=True, color=WHITE, size=9)
        cell.fill      = PatternFill("solid", start_color=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(bottom=Side(style="thin", color="FFAAAAAA"),
                                right=Side(style="thin", color="FFAAAAAA"))

    def dat(cell, val, bg, num_fmt=None):
        cell.value = val
        cell.font      = Font(name=FN, size=8)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(vertical="center")
        cell.border    = Border(bottom=Side(style="hair", color="FFCCCCCC"),
                                right=Side(style="hair", color="FFCCCCCC"))
        if num_fmt: cell.number_format = num_fmt

    def set_widths(ws, ww):
        for i, w in enumerate(ww, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Matched"
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 32
    h1 = ["Customer Email","Customer Name","External ID / Payment ID",
          "Currency","Amount","Date","Bank TRN","Bank Amount","Bank Date"]
    for c, h in enumerate(h1, 1): hdr(ws1.cell(1, c), h)
    set_widths(ws1, [28, 22, 38, 10, 13, 14, 24, 13, 14])

    for ri, (_, brow) in enumerate(matched.iterrows()):
        r = ri + 2
        bg = GREEN if ri % 2 == 0 else GREEN2
        dep = bnk_to_dep.get(brow["_bnk"], {})
        def g(col): return "" if not col or str(dep.get(col,"")) in ("nan","") else str(dep.get(col,"")).strip()
        vals = [g(email_col), g(name_col), g(ext_col), g(cur_col),
                to_amount(dep.get(amt_col, 0)) if amt_col else "",
                g(dat_col)[:10] if g(dat_col) else "",
                brow["_bnk"], brow["_credit"], brow["_date"]]
        fmts = [None,None,None,None,"#,##0.00",None,None,"#,##0.00",None]
        for ci,(v,f) in enumerate(zip(vals,fmts),1): dat(ws1.cell(r,ci), v, bg, f)

    ws2 = wb.create_sheet("Not Found in Deposit")
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 32
    h2 = ["Bank Date","Bank TRN","Remitter","Bank Amount"]
    for c, h in enumerate(h2, 1): hdr(ws2.cell(1, c), h)
    set_widths(ws2, [14, 28, 34, 15])

    for ri, (_, row) in enumerate(not_found.iterrows()):
        r = ri + 2
        bg = RED if ri % 2 == 0 else RED2
        remit = "" if row["_remit"] in ("nan","") else row["_remit"]
        vals = [row["_date"], row["_bnk"], remit, row["_credit"]]
        fmts = [None,None,None,"#,##0.00"]
        for ci,(v,f) in enumerate(zip(vals,fmts),1): dat(ws2.cell(r,ci), v, bg, f)

    out = BytesIO(); wb.save(out); out.seek(0)
    return out, len(matched), len(not_found)


@app.post("/nobimatik")
async def api_nobimatik(
    bank_file: UploadFile = File(...),
    tx_file:   UploadFile = File(...),
):
    bank_data    = await bank_file.read()
    deposit_data = await tx_file.read()
    try:
        result, matched, not_found = build_nobimatik_report(bank_data, deposit_data)
    except Exception as e:
        raise HTTPException(400, str(e))
    return StreamingResponse(
        result,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="nobimatik_report.xlsx"',
            "X-Matched":   str(matched),
            "X-Not-Found": str(not_found),
            "Access-Control-Expose-Headers": "X-Matched, X-Not-Found",
        },
    )
